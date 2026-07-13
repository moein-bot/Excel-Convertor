from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .constants import (
    USER_DATA_SHEET,
    DEFAULT_TEMPLATE_SHEET,
)
from .logger import get_logger

logger = get_logger(__name__)


class WorkbookError(Exception):
    """Base workbook exception."""


class WorkbookNotFoundError(WorkbookError):
    """Workbook file does not exist."""


class SheetNotFoundError(WorkbookError):
    """Worksheet does not exist."""


class EmptyWorksheetError(WorkbookError):
    """Worksheet has no rows."""


class IOHandler:

    def __init__(
        self,
        template_path: str | Path,
        user_path: str | Path,
    ):

        self.template_path = Path(template_path)
        self.user_path = Path(user_path)

        self.template_workbook: Workbook | None = None
        self.user_workbook: Workbook | None = None

    # ----------------------------------------------------

    def load(self):

        self.template_workbook = self._load_workbook(
            self.template_path
        )

        self.user_workbook = self._load_workbook(
            self.user_path,
            data_only=True,
        )

        logger.info("Both workbooks loaded successfully.")

    # ----------------------------------------------------

    def _load_workbook(
        self,
        path: Path,
        data_only: bool = False,
    ) -> Workbook:

        if not path.exists():
            raise WorkbookNotFoundError(path)

        logger.info(f"Loading workbook: {path}")

        return load_workbook(
            filename=path,
            data_only=data_only,
        )

    # ----------------------------------------------------

    def get_user_sheet(self) -> Worksheet:

        return self.get_sheet(
            self.user_workbook,
            USER_DATA_SHEET,
        )

    # ----------------------------------------------------

    def get_template_sheet(self) -> Worksheet:

        return self.get_sheet(
            self.template_workbook,
            DEFAULT_TEMPLATE_SHEET,
        )

    # ----------------------------------------------------

    def get_sheet(
        self,
        workbook: Workbook,
        sheet_name: str,
    ) -> Worksheet:

        if sheet_name not in workbook.sheetnames:

            raise SheetNotFoundError(sheet_name)

        return workbook[sheet_name]

    # ----------------------------------------------------

    @staticmethod
    def read_headers(
        sheet: Worksheet,
        header_row: int = 1,
    ) -> Dict[str, int]:

        headers = {}

        for cell in sheet[header_row]:

            if cell.value is None:
                continue

            headers[str(cell.value).strip()] = cell.column

        return headers

    # ----------------------------------------------------

    @staticmethod
    def read_rows(
        sheet: Worksheet,
        header_row: int = 1,
    ) -> List[dict]:

        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) <= header_row:

            raise EmptyWorksheetError(sheet.title)

        headers = rows[header_row - 1]

        result = []

        for row in rows[header_row:]:

            if all(value is None for value in row):
                continue

            item = {}

            for key, value in zip(headers, row):

                item[str(key).strip()] = value

            result.append(item)

        logger.info(
            f"Loaded {len(result)} records from {sheet.title}"
        )

        return result

    # ----------------------------------------------------

    @staticmethod
    def write_value(
        sheet: Worksheet,
        row: int,
        column: int,
        value,
    ):

        sheet.cell(
            row=row,
            column=column,
            value=value,
        )

    # ----------------------------------------------------

    def save(
        self,
        output_path: str | Path,
    ):

        output_path = Path(output_path)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        self.template_workbook.save(output_path)

        logger.info(
            f"Workbook saved to {output_path}"
        )
